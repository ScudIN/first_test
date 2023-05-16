from openpyxl import load_workbook

def replace(path):
    # openpyxl 워크북 정의
    wb = load_workbook(path)
    # 객체 정의 : 워크북 활성화 시트
    ws1 = wb['Sheet1'] # Sheet1 선택

    # 가변값 read
    value1 = ws1['B3'].value
    value2 = ws1['B4'].value
    value3 = ws1['B5'].value

    # TC 생성 시트 마지막 열, 행
    ws2 = wb['Sheet2'] # Sheet2 선택
    column_max = ws2.max_column
    row_max = ws2.max_row

    # 열/행 for loop
    for col_num in range(1, column_max+1):
        for row_num in range(1, row_max+1):

            #tempstr : cell값이 문자열이 아닌 경우 감안
            tempstr = str(ws2.cell(row = row_num, column = col_num).value)
            #replace
            data = tempstr.replace("대상그룹", value1).replace("작업유형", value2).replace("하위조건", value3)
            #빈 셀의 경우 None이라는 문자열 타입이므로 제외 처리
            if data != "None":
                ws2.cell(row = row_num, column = col_num).value = data
    
    # 저장
    wb.save("TC_result.xlsx")

if __name__ == "__main__":
    path = r"C:/Users/sia0926/Desktop/Py"

replace('C:/Users/sia0926/Desktop/Py/value.xlsx')